#!/usr/bin/env python3
"""
tools/xlsx_edit.py
使用 openpyxl 编辑 Excel .xlsx 文件的工具。
支持：读取/写入单元格、读取/批量填充区域、获取工作表列表、添加/删除工作表。
写操作时若文件不存在则自动新建。
"""

import os
import json
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries

# ==================== 工具定义 ====================

tool_def = {
    "type": "function",
    "function": {
        "name": "xlsx_editor",
        "description": "编辑 Excel xlsx 表格：读取/写入单元格、区域，列出/添加/删除工作表。写操作时若文件不存在则自动创建。",
        "parameters": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "文件路径（必须包含 .xlsx 扩展名）"
                },
                "operation": {
                    "type": "string",
                    "enum": [
                        "read_cell",
                        "write_cell",
                        "read_range",
                        "write_range",
                        "list_sheets",
                        "add_sheet",
                        "delete_sheet"
                    ],
                    "description": "要执行的操作"
                },
                "sheet_name": {
                    "type": "string",
                    "description": "工作表名称（除 list_sheets 外，必需）"
                },
                "cell": {
                    "type": "string",
                    "description": "单元格地址，如 'A1'（用于 read_cell / write_cell）"
                },
                "value": {
                    "description": "要写入单元格的值（用于 write_cell）"
                },
                "range": {
                    "type": "string",
                    "description": "区域范围，如 'A1:B10'（用于 read_range / write_range）"
                },
                "data": {
                    "type": "array",
                    "items": {
                        "type": "array"
                    },
                    "description": "二维数组，用于 write_range 填充区域"
                }
            },
            "required": ["file_path", "operation"]
        }
    }
}

# ==================== 内部辅助函数 ====================

def _ensure_dir(file_path: str):
    """确保文件所在目录存在"""
    dirname = os.path.dirname(file_path)
    if dirname and not os.path.exists(dirname):
        os.makedirs(dirname, exist_ok=True)

def _load_or_create_workbook(file_path: str, for_write: bool = False):
    """
    根据文件是否存在加载工作簿。
    若 for_write=True 且文件不存在，则新建空白工作簿。
    否则若文件不存在则抛出 FileNotFoundError。
    """
    if os.path.exists(file_path):
        try:
            return load_workbook(file_path)
        except Exception as e:
            raise Exception(f"无法加载 Excel 文件: {e}")
    else:
        if for_write:
            # 新建空白工作簿，默认包含一个名为 'Sheet' 的工作表
            wb = Workbook()
            # 默认的 Sheet 名称可能是 'Sheet'，保留即可
            return wb
        else:
            raise FileNotFoundError(f"文件不存在: {file_path}")

def _get_worksheet(wb, sheet_name: str, must_exist: bool = True):
    """获取工作表，若必须存在且不存在则抛出异常"""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    elif must_exist:
        raise Exception(f"工作表 '{sheet_name}' 不存在")
    else:
        return None

def _value_to_json_serializable(obj):
    """
    将 openpyxl 可能返回的非 JSON 可序列化对象转换。
    这里使用 json.dumps 的 default=str 处理，所以不需要手动转换，
    但保留此函数以便将来扩展。
    """
    return obj

def _range_to_boundaries(range_str: str):
    """将区域字符串（如 'A1:B10'）转换为 (min_col, min_row, max_col, max_row)"""
    try:
        return range_boundaries(range_str)
    except Exception as e:
        raise Exception(f"无效的区域格式: {range_str}")

# ==================== 操作实现 ====================

def _read_cell(file_path: str, sheet_name: str, cell: str) -> str:
    wb = _load_or_create_workbook(file_path, for_write=False)
    ws = _get_worksheet(wb, sheet_name)
    try:
        value = ws[cell].value
    except Exception as e:
        raise Exception(f"读取单元格 {cell} 失败: {e}")
    # 使用 default=str 处理日期等非序列化对象
    return json.dumps({"value": value}, ensure_ascii=False, default=str)

def _write_cell(file_path: str, sheet_name: str, cell: str, value) -> str:
    wb = _load_or_create_workbook(file_path, for_write=True)
    # 如果工作表不存在，则创建（仅对写操作）
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    try:
        ws[cell] = value
    except Exception as e:
        raise Exception(f"写入单元格 {cell} 失败: {e}")
    _ensure_dir(file_path)
    wb.save(file_path)
    return json.dumps({"status": "success", "message": f"已写入 {cell}"}, ensure_ascii=False)

def _read_range(file_path: str, sheet_name: str, range_str: str) -> str:
    wb = _load_or_create_workbook(file_path, for_write=False)
    ws = _get_worksheet(wb, sheet_name)
    try:
        cells = ws[range_str]
    except Exception as e:
        raise Exception(f"读取区域 {range_str} 失败: {e}")
    # cells 是一个二维元组，每行是一个元组
    data = []
    for row in cells:
        data.append([cell.value for cell in row])
    return json.dumps({"data": data}, ensure_ascii=False, default=str)

def _write_range(file_path: str, sheet_name: str, range_str: str, data: list) -> str:
    if not isinstance(data, list) or not all(isinstance(row, list) for row in data):
        raise Exception("data 必须为二维列表")
    wb = _load_or_create_workbook(file_path, for_write=True)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    try:
        min_col, min_row, max_col, max_row = _range_to_boundaries(range_str)
    except Exception as e:
        raise Exception(f"无效的区域范围: {range_str}") from e

    # 检查数据维度是否与区域一致
    rows_needed = max_row - min_row + 1
    cols_needed = max_col - min_col + 1
    if len(data) != rows_needed or any(len(row) != cols_needed for row in data):
        raise Exception(f"数据维度 ({len(data)}x{len(data[0]) if data else 0}) 与区域 ({rows_needed}x{cols_needed}) 不匹配")

    for i, row_data in enumerate(data):
        for j, cell_value in enumerate(row_data):
            col = min_col + j
            row = min_row + i
            cell = ws.cell(row=row, column=col)
            cell.value = cell_value

    _ensure_dir(file_path)
    wb.save(file_path)
    return json.dumps({"status": "success", "message": f"已填充区域 {range_str}"}, ensure_ascii=False)

def _list_sheets(file_path: str) -> str:
    wb = _load_or_create_workbook(file_path, for_write=False)
    return json.dumps({"sheets": wb.sheetnames}, ensure_ascii=False)

def _add_sheet(file_path: str, sheet_name: str) -> str:
    wb = _load_or_create_workbook(file_path, for_write=True)
    if sheet_name in wb.sheetnames:
        raise Exception(f"工作表 '{sheet_name}' 已存在")
    wb.create_sheet(sheet_name)
    _ensure_dir(file_path)
    wb.save(file_path)
    return json.dumps({"status": "success", "message": f"已添加工作表 '{sheet_name}'"}, ensure_ascii=False)

def _delete_sheet(file_path: str, sheet_name: str) -> str:
    wb = _load_or_create_workbook(file_path, for_write=True)
    if sheet_name not in wb.sheetnames:
        raise Exception(f"工作表 '{sheet_name}' 不存在")
    # 不能删除最后一个工作表，但 openpyxl 允许，删除后工作簿将无工作表，可能引发问题。
    # 这里允许删除，但若删除后无工作表，openpyxl 保存时可能出错。我们可以在删除后检查，
    # 如果工作簿没有工作表，则新建一个默认工作表以避免问题。
    wb.remove(wb[sheet_name])
    if len(wb.sheetnames) == 0:
        wb.create_sheet("Sheet")  # 创建一个默认工作表
    _ensure_dir(file_path)
    wb.save(file_path)
    return json.dumps({"status": "success", "message": f"已删除工作表 '{sheet_name}'"}, ensure_ascii=False)

# ==================== 主执行函数 ====================

def execute(file_path: str, operation: str, **kwargs) -> str:
    """
    执行 Excel 编辑操作。
    参数：
        file_path : str - 文件路径
        operation : str - 操作类型
        其他关键字参数依据 operation 传递。
    返回：
        JSON 格式字符串（成功时）或错误信息字符串。
    """
    try:
        if operation == "read_cell":
            sheet_name = kwargs.get("sheet_name")
            cell = kwargs.get("cell")
            if not sheet_name or not cell:
                raise Exception("read_cell 需要 sheet_name 和 cell 参数")
            return _read_cell(file_path, sheet_name, cell)

        elif operation == "write_cell":
            sheet_name = kwargs.get("sheet_name")
            cell = kwargs.get("cell")
            value = kwargs.get("value")
            if not sheet_name or not cell or value is None:
                raise Exception("write_cell 需要 sheet_name, cell 和 value 参数")
            return _write_cell(file_path, sheet_name, cell, value)

        elif operation == "read_range":
            sheet_name = kwargs.get("sheet_name")
            range_str = kwargs.get("range")
            if not sheet_name or not range_str:
                raise Exception("read_range 需要 sheet_name 和 range 参数")
            return _read_range(file_path, sheet_name, range_str)

        elif operation == "write_range":
            sheet_name = kwargs.get("sheet_name")
            range_str = kwargs.get("range")
            data = kwargs.get("data")
            if not sheet_name or not range_str or data is None:
                raise Exception("write_range 需要 sheet_name, range 和 data 参数")
            return _write_range(file_path, sheet_name, range_str, data)

        elif operation == "list_sheets":
            return _list_sheets(file_path)

        elif operation == "add_sheet":
            sheet_name = kwargs.get("sheet_name")
            if not sheet_name:
                raise Exception("add_sheet 需要 sheet_name 参数")
            return _add_sheet(file_path, sheet_name)

        elif operation == "delete_sheet":
            sheet_name = kwargs.get("sheet_name")
            if not sheet_name:
                raise Exception("delete_sheet 需要 sheet_name 参数")
            return _delete_sheet(file_path, sheet_name)

        else:
            raise Exception(f"未知操作: {operation}")

    except FileNotFoundError as e:
        return json.dumps({"error": f"文件不存在: {e}"}, ensure_ascii=False)
    except Exception as e:
        # 返回错误信息
        return json.dumps({"error": str(e)}, ensure_ascii=False)